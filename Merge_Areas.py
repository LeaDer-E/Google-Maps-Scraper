import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def merge_and_filter(folder_path, output_file):
    """
    1. Reads all Excel files (columns A:F) in the specified folder.
    2. Cleans cell values by removing extra spaces and hidden characters.
    3. Deletes rows where both the Phone Number (column D) and Website (column E) are empty.
    4. Separates rows with and without phone numbers.
    5. Sorts rows with phone numbers by Phone Number and removes duplicate phone numbers (keeping only the first occurrence).
    6. Sorts rows without phone numbers by Website and removes duplicate websites (keeping only the first occurrence).
    7. Merges the results, writes to an Excel file, and applies formatting.
    """
    all_data = []

    # Function to clean and normalize each cell
    def clean_cell(val):
        if pd.isna(val):
            return ''
        val = str(val).strip()
        # Remove potential hidden characters
        val = val.replace('\xa0', '').replace('\u200f', '').replace('\n', '').replace('\t', '').replace('\r', '')
        return val

    # Read all Excel files in the folder
    for file in os.listdir(folder_path):
        if file.endswith(".xlsx") or file.endswith(".xls"):
            file_path = os.path.join(folder_path, file)
            df = pd.read_excel(file_path, usecols="A:F")  # Read columns A to F

            # Clean data
            df = df.applymap(clean_cell)

            # Ensure consistent column names
            df.columns = ["Place Name", "Description", "Address", "Phone Number", "Website", "Area"]

            all_data.append(df)

    if not all_data:
        print("⚠️ No Excel files found in the specified folder.")
        return

    # Merge all data into one DataFrame
    merged_df = pd.concat(all_data, ignore_index=True)

    # Remove rows where both Phone Number and Website are empty
    mask_both_empty = (merged_df["Phone Number"] == "") & (merged_df["Website"] == "")
    merged_df = merged_df[~mask_both_empty].copy()

    # Separate rows with phone numbers and rows without
    df_with_phone = merged_df[merged_df["Phone Number"] != ""].copy()
    df_without_phone = merged_df[merged_df["Phone Number"] == ""].copy()

    # For rows with phone numbers: sort by Phone Number and remove duplicates (keep first)
    df_with_phone = df_with_phone.sort_values(by="Phone Number")
    df_with_phone = df_with_phone.drop_duplicates(subset=["Phone Number"], keep="first")

    # For rows without phone numbers: sort by Website and remove duplicates (keep first)
    df_without_phone = df_without_phone.sort_values(by="Website")
    df_without_phone = df_without_phone.drop_duplicates(subset=["Website"], keep="first")

    # Combine both parts
    final_df = pd.concat([df_with_phone, df_without_phone], ignore_index=True)

    # Save the result to an Excel file
    final_df.to_excel(output_file, index=False)

    # Apply formatting to the output file
    format_excel(output_file)
    print("✅ Excel files merged, filtered, and cleaned successfully.")

def format_excel(filename):
    # Open the workbook for formatting
    wb = load_workbook(filename)
    ws = wb.active

    # Define cell styles
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_font = Font(color="FFFFFF", size=18, bold=True)
    orange_font = Font(color="E28743", size=18, bold=True)
    darkyellow_font = Font(color="E2A336", size=18, bold=True)
    dark_gray_border = Border(
        left=Side(color='808080'),
        right=Side(color='808080'),
        top=Side(color='808080'),
        bottom=Side(color='808080')
    )
    alignment_center = Alignment(horizontal='center', vertical='center')

    # Set column widths for each column
    column_widths = {
        'A': 30.00,  # Place Name
        'B': 35.00,  # Description
        'C': 60.00,  # Address
        'D': 35.00,  # Phone Number
        'E': 40.00,  # Website
        'F': 30.00   # Area
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze panes at cell B2
    ws.freeze_panes = 'B2'

    # Set row height for the header row
    ws.row_dimensions[1].height = 45

    # Apply styles to all cells in the range A:F
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.fill = black_fill
            cell.border = dark_gray_border
            if cell.row == 1:
                # Header row formatting
                cell.font = orange_font
            elif cell.column_letter == 'D' and cell.row > 1:
                # Phone Number column formatting
                cell.font = darkyellow_font
            else:
                cell.font = white_font
            cell.alignment = alignment_center

    wb.save(filename)
    print("✅ Formatting applied successfully.")

# Example usage
folder_path = "ready/"       # Change this to the folder containing the Excel files
output_file = "merged.xlsx"  # Name of the final Excel file
merge_and_filter(folder_path, output_file)
