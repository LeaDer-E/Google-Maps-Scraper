from selenium import webdriver
import time
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
import undetected_chromedriver as uc
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from colorama import init, Fore, Style

# Initialize colorama for cross-platform color support
init(autoreset=True)

# Reduce logging output from undetected_chromedriver
logging.getLogger("undetected_chromedriver").setLevel(logging.CRITICAL)

# Configure Chrome options to avoid automation detection
chrome_options = uc.ChromeOptions()
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--disable-popup-blocking")
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--remote-debugging-port=0")
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
chrome_options.headless = True

# Launch the browser
driver = uc.Chrome(options=chrome_options)

found_elements = 0

def print_info(message: str):
    print(Fore.CYAN + message + Style.RESET_ALL)

def print_success(message: str):
    print(Fore.GREEN + message + Style.RESET_ALL)

def print_error(message: str):
    print(Fore.RED + message + Style.RESET_ALL)

def print_divider(char: str = "-", color=Fore.BLUE):
    print(color + char * 60 + Style.RESET_ALL)

# Function to print the program header
def print_header():
    print_divider()
    print(Fore.YELLOW + "╔" + "═" * 58 + "╗")
    print(Fore.YELLOW + "║" + Style.BRIGHT + " Google Maps Scraper for Multiple Areas ".center(58) + Style.RESET_ALL + Fore.YELLOW + "║")
    print(Fore.YELLOW + "║" + " Extracting place details from Google Maps ".center(58) + Fore.YELLOW + "║")
    print(Fore.YELLOW + "╚" + "═" * 58 + "╝" + Style.RESET_ALL)
    print_divider()
    print()

# Function to scroll and collect places
def scroll_panel(driver):
    print(Fore.CYAN + "Scrolling to collect all places..." + Style.RESET_ALL)
    print_divider()
    global found_elements
    try:
        scrollable_div = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]"))
        )
    except TimeoutException:
        try:
            scrollable_div = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]"))
            )
        except TimeoutException:
            scrollable_div = None

    if scrollable_div:
        start_time = time.time()
        while time.time() - start_time < 30:
            driver.execute_script("arguments[0].scrollTop += 500;", scrollable_div)
            time.sleep(0.5)
        
        elements = scrollable_div.find_elements(By.XPATH, "./div")
        found_elements = len(elements)+1

        print(Fore.GREEN + f"About {int(found_elements/2)} Places Were Found" + Style.RESET_ALL)
        print_divider()
    else:
        print(Fore.RED + "Scrollable element not found with any of the specified XPaths" + Style.RESET_ALL)

    back_to_top()
    time.sleep(5)
    return extract_all()

# Function to go back to the top
def back_to_top():
    try:
        button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/button/div/div[1]"))
        )
        button.click()
        print(Fore.GREEN + "Clicked 'Back to top' button successfully" + Style.RESET_ALL)
        print_divider()
    except TimeoutException:
        print(Fore.RED + "Failed to find or click 'Back to top' button" + Style.RESET_ALL)


'''
# Function to extract place details
def extract_all():
    results = []
    clickable_xpaths = [f"/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div[{i}]" for i in range(1, found_elements)]
    clickable_xpaths1 = [f"/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div[{i}]/div]" for i in range(1, found_elements)]
    clickable_xpaths2 = [f"/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div[{i}]/div/a" for i in range(1, found_elements)]
    
    
    print(Fore.CYAN + "Starting to extract place details..." + Style.RESET_ALL)
    print(Fore.BLUE + "─" * 60 + Style.RESET_ALL)
    for idx, xpath in enumerate(clickable_xpaths, start=1):
        try:
            clickable = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            clickable.click()
            clickable.click()
            time.sleep(1)
            clickable.click()
            time.sleep(3)
            
            data = {}
            try:
                panel_info = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]"))
                )
            except Exception:
                print(Fore.RED + f"Failed to load details panel for element {idx}" + Style.RESET_ALL)
                continue
            
            try:
                name_element = panel_info.find_element(By.XPATH, ".//h1[contains(@class, 'DUwDvf') and contains(@class, 'lfPIob')]//span[@dir='ltr']")
            except Exception:
                try:
                    name_element = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]/div[2]/div/div[1]/div[1]/h1")
                except Exception:
                    name_element = None

            if name_element:
                data["Place Name"] = name_element.text.strip()
                print(Fore.YELLOW + f"Extracted: {data['Place Name']}" + Style.RESET_ALL)
            else:
                data["Place Name"] = "N/A"
                print(Fore.RED + "Place name not found" + Style.RESET_ALL)

            try:
                desc_element = panel_info.find_element(By.XPATH, ".//button[contains(@class, 'DkEaL')]")
                data["Description"] = desc_element.text.strip()
                print(Fore.CYAN + f"Description: {data['Description']}" + Style.RESET_ALL)
            except Exception:
                data["Description"] = "N/A"

            try:
                address_element = panel_info.find_element(By.XPATH, ".//button[@data-item-id='address']")
                aria = address_element.get_attribute("aria-label")
                data["Address"] = aria.split(":", 1)[1].strip() if aria and ":" in aria else address_element.text.strip()
                print(Fore.CYAN + f"Address: {data['Address']}" + Style.RESET_ALL)
            except Exception:
                data["Address"] = "N/A"

            try:
                phone_element = panel_info.find_element(By.XPATH, ".//button[starts-with(@data-item-id, 'phone:tel')]")
                aria = phone_element.get_attribute("aria-label")
                data["Phone Number"] = aria.split(":", 1)[1].strip() if aria and ":" in aria else phone_element.text.strip()
                print(Fore.CYAN + f"Phone Number: {data['Phone Number']}" + Style.RESET_ALL)
            except Exception:
                data["Phone Number"] = "N/A"

            try:
                website_element = panel_info.find_element(By.XPATH, ".//a[@data-item-id='authority']")
                data["Website"] = website_element.get_attribute("href").strip()
                print(Fore.CYAN + f"Website: {data['Website']}" + Style.RESET_ALL)
            except Exception:
                data["Website"] = "N/A"
            
            results.append(data)
            print(Fore.BLUE + "─" * 60 + Style.RESET_ALL)
            time.sleep(1)
        except Exception:
            print(Fore.RED + f"Error extracting details for element {idx}" + Style.RESET_ALL)
            continue
    
    return results

'''

def extract_all():
    results = []
    # Generate three lists of XPaths using found_elements as the upper bound.
    clickable_xpaths = [f"/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div[{i}]" for i in range(1, found_elements)]
    clickable_xpaths1 = [f"/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div[{i}]/div" for i in range(1, found_elements)]
    clickable_xpaths2 = [f"/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div[{i}]/div/a" for i in range(1, found_elements)]
    
    print(Fore.CYAN + "Starting to extract place details..." + Style.RESET_ALL)
    print(Fore.BLUE + "─" * 60 + Style.RESET_ALL)
    
    # Iterate through the three lists in parallel
    for idx, (xpath0, xpath1, xpath2) in enumerate(zip(clickable_xpaths, clickable_xpaths1, clickable_xpaths2), start=1):
        try:
            # Find and scroll the first element into view
            clickable0 = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath0)))
            driver.execute_script("arguments[0].scrollIntoView(true);", clickable0)
            time.sleep(0.5)
            clickable0.click()
            time.sleep(1)
            
            # Find and scroll the second element into view
            clickable1 = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath1)))
            driver.execute_script("arguments[0].scrollIntoView(true);", clickable1)
            time.sleep(0.5)
            clickable1.click()
            time.sleep(1)
            
            # Find and scroll the third element into view
            clickable2 = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath2)))
            driver.execute_script("arguments[0].scrollIntoView(true);", clickable2)
            time.sleep(0.5)
            clickable2.click()
            time.sleep(3)
            
            data = {}
            try:
                panel_info = WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]"))
                )
            except Exception:
                print(Fore.RED + f"Failed to load details panel for element {idx}" + Style.RESET_ALL)
                continue
            
            try:
                name_element = panel_info.find_element(By.XPATH, ".//h1[contains(@class, 'DUwDvf') and contains(@class, 'lfPIob')]//span[@dir='ltr']")
            except Exception:
                try:
                    name_element = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]/div[2]/div/div[1]/div[1]/h1")
                except Exception:
                    name_element = None

            if name_element:
                data["Place Name"] = name_element.text.strip()
                print(Fore.YELLOW + f"Extracted: {data['Place Name']}" + Style.RESET_ALL)
            else:
                data["Place Name"] = "N/A"
                print(Fore.RED + "Place name not found" + Style.RESET_ALL)

            try:
                desc_element = panel_info.find_element(By.XPATH, ".//button[contains(@class, 'DkEaL')]")
                data["Description"] = desc_element.text.strip()
                print(Fore.CYAN + f"Description: {data['Description']}" + Style.RESET_ALL)
            except Exception:
                data["Description"] = "N/A"

            try:
                address_element = panel_info.find_element(By.XPATH, ".//button[@data-item-id='address']")
                aria = address_element.get_attribute("aria-label")
                data["Address"] = aria.split(":", 1)[1].strip() if aria and ":" in aria else address_element.text.strip()
                print(Fore.CYAN + f"Address: {data['Address']}" + Style.RESET_ALL)
            except Exception:
                data["Address"] = "N/A"

            try:
                phone_element = panel_info.find_element(By.XPATH, ".//button[starts-with(@data-item-id, 'phone:tel')]")
                aria = phone_element.get_attribute("aria-label")
                data["Phone Number"] = aria.split(":", 1)[1].strip() if aria and ":" in aria else phone_element.text.strip()
                print(Fore.CYAN + f"Phone Number: {data['Phone Number']}" + Style.RESET_ALL)
            except Exception:
                data["Phone Number"] = "N/A"

            try:
                website_element = panel_info.find_element(By.XPATH, ".//a[@data-item-id='authority']")
                data["Website"] = website_element.get_attribute("href").strip()
                print(Fore.CYAN + f"Website: {data['Website']}" + Style.RESET_ALL)
            except Exception:
                data["Website"] = "N/A"
            
            results.append(data)
            print(Fore.BLUE + "─" * 60 + Style.RESET_ALL)
            time.sleep(1)
        except Exception:
            print(Fore.RED + f"Error extracting details for element {idx}" + Style.RESET_ALL)
            continue

    return results

    '''
    if results:
        today_date = datetime.now().strftime("%m-%d %H-%M-%S")
        filename = f"{search_places} in Multiply Area at {today_date}.xlsx"
        df = pd.DataFrame(results)
        df.to_excel(filename, index=False)

        # Format the Excel file
        format_excel(filename)
        
        print(Fore.GREEN + f"✅ Data extracted and saved to {filename}" + Style.RESET_ALL)
    else:
        print(Fore.RED + "No data was extracted" + Style.RESET_ALL)
    '''
    


# Function to format the Excel file
def format_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    
    # Define styles
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background
    white_font = Font(color="FFFFFF", size=18, bold=True)  # White text
    orange_font = Font(color="E28743", size=18, bold=True)  # Orange text
    darkyellow_font = Font(color="E2A336", size=18, bold=True)  # Dark yellow text
    dark_gray_border = Border(left=Side(color='808080'), right=Side(color='808080'), top=Side(color='808080'), bottom=Side(color='808080'))
    alignment_center = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 30.00
    ws.column_dimensions['B'].width = 35.00
    ws.column_dimensions['C'].width = 60.00
    ws.column_dimensions['D'].width = 35.00
    ws.column_dimensions['E'].width = 40.00
    ws.column_dimensions['F'].width = 30.00  # For Area column
    
    # Freeze panes
    ws.freeze_panes = 'B2'
    
    # Set height for the first row
    ws.row_dimensions[1].height = 45

    # Apply styles to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.fill = black_fill
            cell.border = dark_gray_border
            if cell.row == 1:
                cell.font = orange_font
            elif cell.column_letter == 'D' and cell.row > 1:
                cell.font = darkyellow_font
            else:
                cell.font = white_font
            cell.alignment = alignment_center
    
    wb.save(filename)
    print_success("✅ Excel file formatting is complete.")

# Print the header
print_header()

# Get user input for place to search
search_places = input(Fore.LIGHTCYAN_EX + "Enter the place to search for: " + Style.RESET_ALL)
print()

# Read areas from Areas.txt
with open("Areas.txt", "r") as file:
    areas = [line.strip() for line in file.readlines()]

all_results = []

for area in areas:
    print_divider()
    print(Fore.CYAN + f"Processing area: {area}" + Style.RESET_ALL)
    print_divider()

    # Open Google Maps
    print(Fore.CYAN + "Opening Google Maps..." + Style.RESET_ALL)
    print_divider()
    driver.get("https://www.google.com/maps")
    time.sleep(3)

    # Search for the place in the current area
    search_box = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "q")))
    search_box.send_keys(f"{search_places} in {area}")
    search_box.send_keys(Keys.ENTER)
    print(Fore.CYAN + f"Searching for '{search_places} in {area}'..." + Style.RESET_ALL)
    print_divider()
    time.sleep(20)

    # Scroll and extract data
    results = scroll_panel(driver)
    for result in results:
        result["Area"] = area  # Add area information
    all_results.extend(results)

if all_results:
    today_date = datetime.now().strftime("%m-%d %H-%M-%S")
    filename = f"{search_places} in Multiple Areas at {today_date}.xlsx"

    # Remove duplicates using pandas
    df = pd.DataFrame(all_results).drop_duplicates()

    # Check if there is unique data after removing duplicates
    if not df.empty:
        df.to_excel(filename, index=False)

        # Format the Excel file
        format_excel(filename)
        
        print(Fore.GREEN + f"✅ Data extracted and saved to {filename}" + Style.RESET_ALL)
    else:
        print(Fore.RED + "No unique data extracted after removing duplicates" + Style.RESET_ALL)
else:
    print(Fore.RED + "No data was extracted for any area" + Style.RESET_ALL)
